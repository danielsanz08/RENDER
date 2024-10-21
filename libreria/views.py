from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from .models import Transaccion, Insumo, CustomUser, Backup,Cliente,Producto,Proveedor,LineaTransaccion
from django.contrib.auth.decorators import login_required
from .forms import InsumoForm, TransaccionForm, CustomUserCreationForm, LoginForm, CustomUserChangeForm,ClienteForm,ProductoForm,ProveedorForm
from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth import authenticate, login as auth_login
from .utils import create_backup, restore_backup
import os
from django.http import FileResponse
from .forms import CustomPasswordChangeForm
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.hashers import check_password
from django.contrib.auth import get_user_model
from django.urls import reverse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.cache import never_cache

from google.auth.transport.requests import Request
from django.contrib.staticfiles import finders
import openpyxl
from io import BytesIO
from .models import Proveedor,Insumo, Cliente

from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.colors import Color
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
import openpyxl
import reportlab
import google.auth
@never_cache
def inicio(request):
    usuario = request.user  # Obtiene el usuario actual que ha iniciado sesión
    usuario = request.user
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/inicio'},
    ]
    return render(request, 'index/index.html', {'usuario': usuario, 'breadcrumbs': breadcrumbs})


def usuario(request):
    usuario = request.user  # Obtiene el usuario actual que ha iniciado sesión
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/'},
        {'name': 'Usuario', 'url': '/'},
    ]
    return render(request, 'usuario/usuario.html', {'usuario': usuario,'breadcrumbs': breadcrumbs})

@login_required
def manual_usuario(request):
    return render(request, 'manual/manual.html')

def abrir_pdf(request):
    # Asegúrate de que esta ruta sea correcta para tu proyecto
    filepath = os.path.join('static', 'pdfs', 'manual_usuario.pdf')
    return FileResponse(open(filepath, 'rb'), content_type='application/pdf')

@login_required
def contabilidad(request):
    usuario = request.user  # Obtiene el usuario actual que ha iniciado sesión
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/inicio'},
        {'name': 'Movimientos', 'url': '/contabilidad'},
    ]
    return render(request, 'movimientos/movimientos.html', {'usuario': usuario,'breadcrumbs': breadcrumbs})

@login_required
def insumos(request):
    usuario = request.user  # Obtiene el usuario actual que ha iniciado sesión
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/inicio'},
        {'name': 'Gestión de Proveedores', 'url': '/insumos'},
    ]
    return render(request, 'insumos/insumos.html',{'usuario': usuario, 'breadcrumbs': breadcrumbs})

@login_required
def ver_perfil(request):
    usuario = request.user  # Obtiene el usuario actual que ha iniciado sesión

    if usuario.role == 'Empleado':
        if request.method == 'POST':
            admin_password = request.POST.get('admin_password')
            try:
                admin_user = CustomUser.objects.get(role='Administrador')
                if check_password(admin_password, admin_user.password):
                    return JsonResponse({'success': True})
                else:
                    return JsonResponse({'success': False, 'message': 'Contraseña incorrecta.'})
            except CustomUser.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'No hay un usuario administrador registrado.'})
        
        breadcrumbs = [
            {'name': 'Inicio', 'url': 'Inicio'},
            {'name': 'Usuario', 'url': 'usuario'},
            {'name': 'Ver Perfil', 'url': '/'},
        ]

        return render(request, 'usuario/ver.html', {'usuario': usuario, 'breadcrumbs': breadcrumbs})

    # Si el usuario no es un empleado, simplemente mostrar la página del perfil
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/inicio/'},
        {'name': 'Usuario', 'url': '/usuario/'},
        {'name': 'Ver Perfil', 'url': '/ver_perfil'},
    ]
    return render(request, 'usuario/ver.html', {'usuario': usuario, 'breadcrumbs': breadcrumbs})

@never_cache
@login_required
def editar_perfil(request, user_id):
    usuario = get_object_or_404(CustomUser, id=user_id)  # Obtiene el usuario por ID
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/inicio/'},
        {'name': 'Usuario', 'url': '/usuario/'},
        {'name': 'Usuarios Registrados', 'url': '/listar_usuario/'},
        {'name': 'Editar Perfil', 'url': '/'},
    ]

    # Get the count of existing admins excluding the current user
    admin_users_count = CustomUser.objects.filter(role='Administrador').exclude(id=user_id).count()

    if request.method == 'POST':
        form = CustomUserChangeForm(request.POST, request.FILES, instance=usuario)
        if form.is_valid():
            selected_role = form.cleaned_data.get('role')  # Get the newly selected role
            current_role = usuario.role  # Get the current role

            # Check if the user is changing from Empleado to Administrador
            if current_role == 'Empleado' and selected_role == 'Administrador' and admin_users_count > 0:
                messages.error(request, 'No puedes cambiar el rol de Empleado a Administrador. Ya hay un Administrador registrado.')
                return redirect('editar_perfil', user_id=user_id)  # Redirect to the same page

            # Save the form if validation passed
            form.save()
            messages.success(request, 'Perfil actualizado exitosamente.')
            return redirect('listar_usuario')
    else:
        form = CustomUserChangeForm(instance=usuario)

    return render(request, 'usuario/editar.html', {
        'usuario': usuario,
        'form': form,
        'breadcrumbs': breadcrumbs,
        'admin_users_count': admin_users_count,  # Pass the count to the template if needed
    })
@login_required
def cambiar_contraseña(request):
    usuario = request.user  # Obtiene el usuario actual que ha iniciado sesión
    if request.method == 'POST':
        form = CustomPasswordChangeForm(request.POST, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Tu contraseña ha sido cambiada con éxito. Puedes iniciar sesión con tu nueva contraseña.')
            return redirect('login')  # Redirige a la página de inicio de sesión
        else:
            messages.error(request, 'Hubo un error al cambiar tu contraseña. Asegúrate de que la contraseña actual sea correcta y que las nuevas contraseñas coincidan.')
    else:
        form = CustomPasswordChangeForm(user=request.user)
    
    return render(request, 'usuario/usuario.html', {'usuario': usuario, 'form': form})

@login_required
def listar_usuario(request):
    usuario = request.user  # Obtiene el usuario actual que ha iniciado sesión
    usuarios = CustomUser.objects.all()  # Filtra por rol 'Empleado'
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/inicio/'},
        {'name': 'Usuario', 'url': '/usuario/'},
        {'name': 'Usuarios Registrados', 'url': '/'},
    ]
    return render(request, 'usuario/listar.html', {'usuario': usuario, 'usuarios': usuarios, 'breadcrumbs': breadcrumbs})

def login_view(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            name = form.cleaned_data.get('name')
            password = form.cleaned_data.get('password')
            user = authenticate(request, name=name, password=password)

            if user is not None:
                if user.is_active:  # Verificar si el usuario está activo
                    auth_login(request, user)
                    return redirect('inicio')  # Redirige a la página principal o a otra página
                else:
                    messages.error(request, 'Tu cuenta está inactiva. Contacta al administrador.')
            else:
                messages.error(request, 'Nombre o contraseña incorrectos.')
    else:
        form = LoginForm()

    # Verificar si existe al menos un administrador en la base de datos
    admin_exists = CustomUser.objects.filter(role='Administrador').exists()

    return render(request, 'registration/login.html', {'form': form, 'admin_exists': admin_exists})
def logout_view(request):
    logout(request)
    messages.success(request, 'Has cerrado sesión exitosamente.')
    return redirect('login')  # Redirige a la página de inicio de sesión u otra página deseada

def validar_datos(request):
    if request.method == "GET":
        nombre = request.GET.get('name', None)
        email = request.GET.get('email', None)
        telefono = request.GET.get('phone', None)
        password = request.GET.get('password', None)

        # Validar si ya existe un usuario con el mismo nombre, email o teléfono
        if CustomUser.objects.filter(name=nombre).exists():
            return JsonResponse({'error': 'El nombre ya está en uso.'}, status=400)
        if CustomUser.objects.filter(email=email).exists():
            return JsonResponse({'error': 'El email ya está en uso.'}, status=400)
        if CustomUser.objects.filter(phone=telefono).exists():
            return JsonResponse({'error': 'El teléfono ya está en uso.'}, status=400)
        if CustomUser.objects.filter(password=password).exists():
            return JsonResponse({'error': 'La contraseña ya está en uso.'}, status=400)

        # Si todo está bien
        return JsonResponse({'valid': True})

    return JsonResponse({'error': 'Método no permitido.'}, status=405)

def crear_perfil(request):
    usuario = request.user  # Obtiene el usuario actual que ha iniciado sesión
    form = CustomUserCreationForm()  # Inicializamos el formulario por defecto al inicio

    if request.method == 'POST':
        # Si el usuario es empleado, necesita la contraseña del administrador
        if hasattr(request.user, 'role') and request.user.role == 'Empleado':
            admin_password = request.POST.get('admin_password')
            try:
                # Usa CustomUser en lugar de User
                admin_user = CustomUser.objects.get(role='Administrador')
                if check_password(admin_password, admin_user.password):
                    form = CustomUserCreationForm(request.POST, request.FILES)
                    if form.is_valid():
                        user = form.save()
                        messages.success(request, f'Usuario {user.email} creado exitosamente.')
                        return redirect('login')
                    else:
                        messages.error(request, 'Por favor corrige los errores a continuación.')
                else:
                    messages.error(request, 'Contraseña del administrador incorrecta.')
            except CustomUser.DoesNotExist:  # Cambia esto también a CustomUser
                messages.error(request, 'No hay un usuario administrador registrado.')
        else:
            # Si el usuario es administrador, verificar si ya existe uno antes de guardar
            form = CustomUserCreationForm(request.POST, request.FILES)
            if form.is_valid():
                if form.cleaned_data['role'] == 'Administrador' and CustomUser.objects.filter(role='Administrador').exists():
                    messages.error(request, 'Ya existe un administrador registrado.')
                else:
                    user = form.save()
                    messages.success(request, f'Usuario {user.email} creado exitosamente.')
                    return redirect('login')
            else:
                messages.error(request, 'Por favor corrige los errores a continuación.')

    breadcrumbs = [
        {'name': 'Inicio', 'url': '/'},
        {'name': 'Usuario', 'url': '/usuario/'},
        {'name': 'Crear Perfil', 'url': '/crear_perfil'},
    ]

    # Safely access role
    user_role = getattr(request.user, 'role', None)  # Defaults to None if no role exists

    return render(request, 'usuario/registro.html', {'usuario': usuario, 'form': form, 'breadcrumbs': breadcrumbs, 'user_role': user_role})
@login_required

def crear_transacciones(request):
    usuario = request.user  # Obtiene el usuario actual que ha iniciado sesión

    if request.method == 'POST':
        form = TransaccionForm(request.POST)
        
        if form.is_valid():
            nueva_transaccion = form.save(commit=False)
            nueva_transaccion.registrado_por = usuario  # Asignar el usuario actual
            
            # Asegúrate de obtener el ID del producto desde el formulario
            producto_id = request.POST.get('producto')  # Suponiendo que tienes un campo de producto
            producto = get_object_or_404(Producto, pk=producto_id)
            nueva_transaccion.producto = producto  # Asignar el producto a la transacción
            
            nueva_transaccion.save()  # Guardar la transacción

            # Guardar las líneas de transacción, si es necesario
            cantidades = request.POST.getlist('cantidades[]')  # Ajusta si necesitas manejar múltiples cantidades

            for cantidad in cantidades:
                LineaTransaccion.objects.create(
                    transaccion=nueva_transaccion,
                    producto=producto,
                    cantidad=cantidad  # Si solo hay un producto, esta cantidad debe ser la misma
                )

            messages.success(request, 'Transacción creada exitosamente.')
            return redirect('ver_transacciones')
        else:
            messages.error(request, 'Por favor corrige los errores a continuación.')
            print(form.errors)  # Mostrar errores de validación

    else:
        form = TransaccionForm()

    clientes = Cliente.objects.all()
    productos = Producto.objects.all()

    breadcrumbs = [
        {'name': 'Inicio', 'url': '/'},
        {'name': 'Movimientos', 'url': '/contabilidad'},
        {'name': 'Agregar Movimiento', 'url': '/crear_transacciones'},
    ]

    return render(request, 'movimientos/crear_movimiento.html', {
        'usuario': usuario,
        'form': form,
        'clientes': clientes,
        'productos': productos,
        'breadcrumbs': breadcrumbs,
    })

def ver_transacciones(request):
    usuario = request.user
    transacciones = Transaccion.objects.select_related('producto', 'cliente').all()

    breadcrumbs = [
        {'name': 'Inicio', 'url': '/'},
        {'name': 'Movimientos', 'url': '/contabilidad'},
        {'name': 'Movimientos Registrados', 'url': '/ver_transacciones'},
    ]

    return render(request, 'movimientos/ver_movimientos.html', {
        'usuario': usuario,
        'transacciones': transacciones,
        'breadcrumbs': breadcrumbs
    })



@login_required
def registros_recientes(request):
    recientes = Transaccion.objects.all().order_by('-fecha')[:3]
    return render(request, 'movimientos/movimientos.html', {'recientes': recientes})

@login_required
def eliminar(request, id):
    transaccion = Transaccion.objects.get(id=id)
    transaccion.delete()
    return redirect('ver_transacciones')

@login_required
def search(request):
    search_type = request.GET.get('type')
    query = request.GET.get('query')

    if search_type == 'fecha':
        transacciones = Transaccion.objects.filter(fecha__icontains=query)
    elif search_type == 'descripcion':
        transacciones = Transaccion.objects.filter(descripcion__icontains=query)
    else:
        transacciones = Transaccion.objects.all()

    results = list(transacciones.values('tipo', 'descripcion', 'monto', 'fecha'))
    return JsonResponse(results, safe=False)

@login_required
def agregar_insumo(request):
    usuario = request.user  # Obtiene el usuario actual que ha iniciado sesión
    if request.method == 'POST':
        form = InsumoForm(request.POST)
        if form.is_valid():
            insumo = form.save(commit=False)
            proveedor_id = request.POST.get('proveedor')  # Captura el proveedor seleccionado
            try:
                proveedor = Proveedor.objects.get(id=proveedor_id)  # Asegúrate de que existe el proveedor
                insumo.proveedor = proveedor  # Asocia el proveedor al insumo
                insumo.registrado_por = request.user  # Asigna el usuario que registra
                insumo.save()
                messages.success(request, 'Insumo agregado exitosamente.')
                return redirect('consultar_insumo')
            except Proveedor.DoesNotExist:
                messages.error(request, 'Proveedor no válido.')
        else:
            messages.error(request, 'Por favor corrige los errores a continuación.')
    else:
        form = InsumoForm()

    proveedores = Proveedor.objects.all()  # Trae todos los proveedores de la base de datos

    breadcrumbs = [
        {'name': 'Inicio', 'url': '/inicio'},
        {'name': 'Gestión de Proveedores', 'url': '/insumos'},
        {'name': 'Agregar insumos', 'url': '/agregar_insumo'},
    ]

    return render(request, 'insumos/agregar_insumo.html', {
        'usuario': usuario, 
        'form': form,
        'proveedores': proveedores,
        'breadcrumbs': breadcrumbs,
    })


@login_required
def editar_insumo(request, insumo_id):
    usuario = request.user  # Obtiene el usuario actual que ha iniciado sesión
    insumo = Insumo.objects.get(id=insumo_id)

    if request.method == 'POST':
        form = InsumoForm(request.POST, instance=insumo)
        if form.is_valid():
            form.save()
            messages.success(request, 'Insumo actualizado exitosamente.')
            return redirect('consultar_insumo')
        else:
            messages.error(request, 'Por favor corrige los errores a continuación.')
    else:
        form = InsumoForm(instance=insumo)

    return render(request, 'insumos/editar_insumo.html', {
        'usuario': usuario, 
        'form': form,
    })


@login_required
def verificar_nombre_insumo(request):
    nombre = request.GET.get('nombre', None)
    exists = Insumo.objects.filter(nombre=nombre).exists()
    return JsonResponse({'exists': exists})


def verificar_administrador(request):
    # Verificar si hay un usuario con el rol de Administrador
    existe_administrador = CustomUser.objects.filter(role='Administrador').exists()
    
    # Devolver la respuesta en formato JSON
    return JsonResponse({'existe_administrador': existe_administrador})

@login_required
def consultar_insumo(request):
    usuario = request.user  # Obtiene el usuario actual que ha iniciado sesión
    query = request.GET.get('q', '')
    if query:
        insumos = Insumo.objects.filter(nombre__icontains=query)
    else:
        insumos = Insumo.objects.all()

    breadcrumbs = [
        {'name': 'Inicio', 'url': '/inicio'},
        {'name': 'Gestión de Proveedores', 'url': '/insumos'},
        {'name': 'Insumos Registrados', 'url': '/consultar_insumo'},
    ]

    return render(request, 'insumos/consultar_insumo.html', {'usuario': usuario, 'insumos': insumos, 'breadcrumbs': breadcrumbs})

def backup_view(request):
    if request.method == 'POST':
        if 'create_backup' in request.POST:
            backup = create_backup()
            if backup:
                messages.success(request, 'Copia de seguridad creada exitosamente.')
            else:
                messages.error(request, 'Error al crear la copia de seguridad.')
        elif 'restore_backup' in request.POST:
            backup_id = request.POST.get('backup_id')
            if restore_backup(backup_id):
                messages.success(request, 'Base de datos restaurada exitosamente.')
            else:
                messages.error(request, 'Error al restaurar la base de datos.')
        elif 'delete_backup' in request.POST:
            backup_id = request.POST.get('backup_id')
            backup = Backup.objects.get(id=backup_id)
            backup.delete()
            messages.success(request, 'Copia de seguridad eliminada exitosamente.')

    backups = Backup.objects.all().order_by('-created_at')
    return render(request, 'backup_restore/backup.html', {'backups': backups})

def download_backup(request, id):
    backup = get_object_or_404(Backup, id=id)
    response = HttpResponse(backup.file, content_type='application/force-download')
    response['Content-Disposition'] = f'attachment; filename="{backup.file.name}"'
    return response


def cambiar_estado_usuario(request, user_id):
    if request.method == 'POST':
        user = get_object_or_404(CustomUser, id=user_id)
        user.is_active = 'is_active' in request.POST
        user.save()
        return redirect('login')  # Redirige a la página que quiera
    
    
def clientes(request):
    usuario = request.user  # Obtiene el usuario actual que ha iniciado sesión
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/inicio/'},
        {'name': 'Clientes', 'url': '/clientes/'}
    ]
    return render(request, 'clientes/clientes.html', {'usuario': usuario, 'breadcrumbs': breadcrumbs})

def consultar_clientes(request): 
    usuario = request.user  # Obtiene el usuario actual que ha iniciado sesión
    clientes = Cliente.objects.all()  # Obtener todos los productos

    breadcrumbs = [
        {'name': 'Inicio', 'url': '/inicio/'},
        {'name': 'Clientes', 'url': '/clientes/'},
        {'name': 'Clientes Registrados', 'url': '/'}
    ]

    return render(request, 'clientes/consultar_clientes.html', {'usuario': usuario, 'clientes': clientes, 'breadcrumbs': breadcrumbs})
    
def crear_cliente(request):
    if request.method == 'POST':
        cliente_form = ClienteForm(request.POST)  # Crea el formulario con los datos POST
        if cliente_form.is_valid():  # Verifica si el formulario es válido
            cliente = cliente_form.save(commit=False)  
            cliente.registrado_por = request.user  # Asigna el usuario que registra
            cliente.save()  # Luego, guarda el cliente en la base de datos
            messages.success(request, 'Cliente creado con éxito.')  # Mensaje de éxito
            return redirect('consultar_clientes')  # Redirige a la lista de clientes
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')  # Mensaje de error
            print(cliente_form.errors)  # Imprime los errores del formulario en la consola
    else:
        cliente_form = ClienteForm()  # Si no es POST, crea un formulario vacío

    breadcrumbs = [
        {'name': 'Inicio', 'url': '/inicio/'},
        {'name': 'Clientes', 'url': '/clientes/'},
        {'name': 'Agregar Cliente', 'url': '/'}  # Asegúrate de que esta URL sea correcta
    ]

    return render(request, 'clientes/crear_cliente.html', {
        'usuario': request.user,  # Proporciona el usuario actual
        'cliente_form': cliente_form, 
        'breadcrumbs': breadcrumbs
    })
@login_required
def verificar_cliente(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre', None)
        email = request.POST.get('email', None)
        telefono = request.POST.get('telefono', None)

        # Inicializa la variable de existencia
        exists = False
        field = ''

        # Verifica si ya existe un cliente con el mismo nombre, email o teléfono
        if nombre and Cliente.objects.filter(nombre=nombre).exists():
            exists = True
            field = 'nombre'
        elif email and Cliente.objects.filter(email=email).exists():
            exists = True
            field = 'email'
        elif telefono and Cliente.objects.filter(telefono=telefono).exists():
            exists = True
            field = 'telefono'

        # Devuelve una respuesta JSON con el resultado
        return JsonResponse({'exists': exists, 'field': field})

    return JsonResponse({'exists': False, 'field': ''})

def editar_cliente(request, pk):
    usuario = request.user  # Obtiene el usuario actual que ha iniciado sesión
    cliente = get_object_or_404(Cliente, pk=pk)

    if request.method == 'POST':
        form = ClienteForm(request.POST, request.FILES, instance=cliente)  # Mantén el instance para la actualización
        if form.is_valid():
            form.save()  # Solo se guardarán los campos que han sido modificados
            messages.success(request, 'Cliente actualizado exitosamente.')
            return redirect('consultar_clientes')
        else:
            messages.error(request, 'Error al actualizar el cliente. Por favor, revisa los datos ingresados.')

    else:
        form = ClienteForm(instance=cliente)  # Muestra el formulario con los datos existentes

    breadcrumbs = [
        {'name': 'Inicio', 'url': '/inicio/'},
        {'name': 'Clientes', 'url': '/clientes/'},
        {'name': 'Clientes Registrados', 'url': '/consultar_clientes/'},
        {'name': 'Editar Cliente', 'url': '/'}  # Aquí puedes cambiar la URL si es necesario
    ]

    return render(request, 'clientes/editar_cliente.html', {'form': form, 'cliente': cliente, 'breadcrumbs': breadcrumbs})

    return render(request, 'clientes/editar_cliente.html', {'usuario': usuario, 'form': form, 'cliente': cliente, 'breadcrumbs': breadcrumbs})
def eliminar_cliente(request, pk):
    usuario = request.user
    cliente = get_object_or_404(Cliente, pk=pk)
    
    if request.method == 'POST':
        cliente.delete()
        messages.success(request, 'Cliente eliminado exitosamente.')
        return redirect('consultar_clientes')
    
    return render(request, 'clientes/consultar_clientes.html', {
        'usuario': usuario,
        'cliente': cliente
    })


# Vistas de productos
@login_required
def productos(request):
    usuario = request.user  # Obtiene el usuario actual que ha iniciado sesión
    productos = Producto.objects.all()
    breadcrumbs = [{'name': 'Inicio', 'url': '/'}, {'name': 'Gestión de Productos', 'url': '/'}]
    return render(request, 'productos/producto.html', {'usuario': usuario, 'productos': productos, 'breadcrumbs': breadcrumbs})

def crear_producto(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST)
        if form.is_valid():
            producto = form.save(commit=False)
            # Asegúrate de que temperatura_conservacion sea un número decimal
            producto.temperatura_conservacion = form.cleaned_data['temperatura_conservacion']
            # Asigna el usuario que registra
            producto.registrado_por = request.user  
            producto.save()
            return redirect('lista_productos')
    else:
        form = ProductoForm()
    
    return render(request, 'productos/crear_producto.html', {'form': form})

@login_required
@login_required
def editar_producto(request, pk=None):
    if pk:
        producto = get_object_or_404(Producto, pk=pk)
    else:
        producto = None

    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            producto = form.save(commit=False)
            producto.registrado_por = request.user  # Asignar el usuario que registra
            producto.save()
            messages.success(request, 'El producto ha sido guardado exitosamente.')
            return redirect('lista_productos')
    else:
        form = ProductoForm(instance=producto)

    return render(request, 'productos/editar_producto.html', {
        'form': form,
        'producto': producto
    })



@login_required
def eliminar_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    producto.delete()
    messages.success(request, 'Producto eliminado exitosamente.')
    return redirect('lista_productos')

def lista_productos(request):
    usuario = request.user  # Obtiene el usuario actual que ha iniciado sesión
    productos = Producto.objects.all()  # Obtener todos los productos

    breadcrumbs = [
        {'name': 'Inicio', 'url': '/'},
        {'name': 'Gestión de Productos', 'url': '/productos/'}, 
        {'name': 'Productos Registrados', 'url': '/'}
    ]
    
    return render(request, 'productos/lista_producto.html', {'usuario': usuario, 'productos': productos, 'breadcrumbs': breadcrumbs})


@never_cache
def crear_proveedor(request):
    usuario = request.user  # Obtiene el usuario actual que ha iniciado sesión
    if request.method == 'POST':
        form = ProveedorForm(request.POST)
        if form.is_valid():
            proveedor = form.save(commit=False)
            proveedor.registrado_por = request.user
            proveedor.save()
            return redirect('listar_proveedor')  # Redirige a la tabla de proveedores
        else:
            print(form.errors)  # Mostrar errores del formulario

    else:
        form = ProveedorForm()
        
    # Definimos los breadcrumbs
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/'},
        {'name': 'Gestión de Proveedores', 'url': '/insumos'},
        {'name': 'Crear Proveedor', 'url': '/'}  # No se requiere URL en la página actual
    ]
    
    return render(request, 'proveedores/crear_proveedor.html', {'usuario': usuario, 'form': form, 'breadcrumbs': breadcrumbs})

@never_cache
def listar_proveedor(request):
    usuario = request.user  # Obtiene el usuario actual que ha iniciado sesión
    proveedores = Proveedor.objects.select_related('registrado_por').all()
    breadcrumbs = [
            {'name': 'Inicio', 'url': '/'},
            {'name': 'Gestión de Proveedores', 'url': '/insumos'},
            {'name': 'Proveedores Registrados', 'url': '/listar_proveedor'}  # Para la página actual no es necesario un URL
        ]
    return render(request, 'proveedores/listar_proveedor.html', {'usuario': usuario, 'proveedores': proveedores,'breadcrumbs': breadcrumbs})

@never_cache

def editar_proveedor(request, proveedor_id, campo=None):
    usuario = request.user  # Obtiene el usuario actual que ha iniciado sesión
    proveedor = get_object_or_404(Proveedor, id=proveedor_id)

    # Inicializa el formulario con los datos actuales del proveedor
    form = ProveedorForm(instance=proveedor)

    if request.method == 'POST':
        # Maneja la actualización del campo específico
        if campo in form.fields:
            form = ProveedorForm(request.POST, instance=proveedor)

            if form.is_valid():
                # Actualiza solo el campo específico
                setattr(proveedor, campo, form.cleaned_data[campo])
                proveedor.save()
                messages.success(request, f'{campo.capitalize()} editado exitosamente.')
                return redirect('listar_proveedor')
        
        # Si no es válido, muestra error
        messages.error(request, 'Error al editar el proveedor. Por favor, revisa los errores.')

    breadcrumbs = [
        {'name': 'Inicio', 'url': '/'},
        {'name': 'Gestión de Proveedores', 'url': '/insumos/'},
        {'name': 'Proveedores Registrados', 'url': '/listar_proveedor/'},
        {'name': 'Editar Proveedor', 'url': '#'}
    ]

    return render(request, 'proveedores/editar_proveedor.html', {
        'usuario': usuario,
        'form': form,
        'proveedor': proveedor,
        'breadcrumbs': breadcrumbs,
        'campo': campo  # Envía el campo actual que se está editando
    })
@never_cache
def eliminar_proveedor(request, proveedor_id):
    proveedor = get_object_or_404(Proveedor, id=proveedor_id)
    proveedor.delete()
    messages.success(request, 'Proveedor eliminado exitosamente.')  # Mensaje de éxito
    return redirect('listar_proveedor')  # Asegúrate de que la URL esté configurada

@require_POST
@csrf_exempt
@never_cache
def verificar_proveedor(request):
    if request.method == 'POST':
        nit = request.POST.get('nit')
        nombre = request.POST.get('nombre')
        email = request.POST.get('email')
        telefono = request.POST.get('telefono')

        # Verificar si el proveedor ya existe
        if Proveedor.objects.filter(nit=nit).exists():
            return JsonResponse({'exists': True, 'field': 'nit'})
        if Proveedor.objects.filter(nombre=nombre).exists():
            return JsonResponse({'exists': True, 'field': 'nombre'})
        if Proveedor.objects.filter(email=email).exists():
            return JsonResponse({'exists': True, 'field': 'email'})
        if Proveedor.objects.filter(telefono=telefono).exists():
            return JsonResponse({'exists': True, 'field': 'telefono'})

        # Si no existe, retornamos que no existe
        return JsonResponse({'exists': False})

    return JsonResponse({'exists': False})

def reporte_proveedor_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proveedores"

    logo_path = finders.find('imagen/logo.png')

    # Aumentar el tamaño de la imagen
    img = Image(logo_path)
    img.height = 60  # Ajustar altura
    img.width = 80    # Ajustar ancho
    ws.add_image(img, 'A1')

    ws.merge_cells('B1:C1')
    ws['B1'] = "LACTEOS HEDYBED"
    ws['B1'].font = Font(size=24, bold=True)
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:C2')  # Cambiar a B2:C2 para que abarque correctamente
    ws['A2'] = "Proveedores Registrados"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    headers = ["NIT", "Nombre", "Dirección", "Email", "Teléfono"]
    ws.append(headers)

    # Corregir el color del encabezado a un valor hex válido
    header_fill = PatternFill(start_color="01AB7B", end_color="01AB7B", fill_type="solid")
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)  # Cambiar a blanco para mejor contraste
        cell.alignment = Alignment(horizontal='center', vertical='center')

    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    proveedores = Proveedor.objects.all()
    for proveedor in proveedores:
        nit_formateado = proveedor.formatear_nit()
        ws.append([
            nit_formateado,
            proveedor.nombre,
            proveedor.direccion,
            proveedor.email,
            proveedor.telefono if proveedor.telefono else "N/A",
        ])

    # Establecer el ancho de las columnas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15

    # Aplicar estilos a las celdas
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

    # Ajustar el ancho de la fila para el título
    ws.row_dimensions[1].height = 40  # Ajusta la altura de la fila para que el título se vea bien
    ws.row_dimensions[2].height = 30  # Ajusta la altura de la fila para el subtítulo

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_proveedores.xlsx"'
    
    wb.save(response)
    return response


def reporte_proveedor_pdf(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=landscape(letter))  # Cambiar a landscape
    width, height = landscape(letter)  # Ajustar el ancho y alto a la orientación

    margin = 40
    table_width = width - 2 * margin
    y_position = height - margin - 110  # Ajusta el margen superior para dejar espacio para el encabezado

    # Ruta de la marca de agua
    watermark_path = finders.find('imagen/logo.png')
    p.saveState()

    # Ajustar transparencia de la marca de agua
    p.setFillColor(colors.Color(1, 1, 1, alpha=0.3))
    p.setStrokeColor(colors.Color(1, 1, 1, alpha=0.3))
    p.drawImage(watermark_path, x=(width - 400) / 2, y=(height - 400) / 2, width=400, height=400, mask='auto')

    p.restoreState()

    # Dibujar encabezado
    p.setFont("Helvetica-Bold", 24)
    p.drawCentredString(width / 2, y_position + 50, "LACTEOS HEDYBED")
    p.setFont("Helvetica", 18)
    p.drawCentredString(width / 2, y_position + 20, "Reporte de Proveedores")

    # Crear la tabla de datos
    data = [["NIT", "Nombre", "Dirección", "Email", "Teléfono", "Registrado por"]]
    proveedores = Proveedor.objects.all()
    for proveedor in proveedores:
        nit_formateado = proveedor.formatear_nit()
        registrado_por = proveedor.registrado_por.get_username() if proveedor.registrado_por else 'N/A'
        
        data.append([
            nit_formateado,
            proveedor.nombre,
            proveedor.direccion,
            proveedor.email,
            proveedor.telefono if proveedor.telefono else 'N/A',
            registrado_por
        ])

    # Crear la tabla con los datos y especificar los anchos de columna
    table = Table(data, colWidths=[1 * inch, 1 * inch, 1 * inch, 3 * inch, 1.5 * inch, 2 * inch])  # Anchos ajustados

    # Estilo de la tabla
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#01AB7B")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('TOPPADDING', (0, 1), (-1, -1), 2),  # Espaciado ajustado
        ('BOTTOMPADDING', (0, 1), (-1, -1), 2),  # Espaciado ajustado
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    table.setStyle(style)

    # Ajustar y dibujar la tabla
    table.wrapOn(p, table_width, height - 2 * margin)
    table_x = margin
    table_y = y_position - len(data) * 30  # Ajusta la posición de la tabla
    table.drawOn(p, table_x, table_y)

    # Finalizar la página del PDF
    p.showPage()
    p.save()
    
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte_proveedor.pdf"'

    return response


def reporte_insumo_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Insumos"

    logo_path = finders.find('imagen/logo.png')
    img = Image(logo_path)
    img.height = 60
    img.width = 80
    ws.add_image(img, 'A1')

    ws.merge_cells('B1:D1')
    ws['B1'] = "LACTEOS HEDYBED"
    ws['B1'].font = Font(size=24, bold=True)
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:E2')
    ws['A2'] = "Insumos Registrados"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    headers = ["ID", "Nombre", "Descripción", "Proveedor", "Cantidad"]
    ws.append(headers)

    # Corregir el color del encabezado a un valor hex válido
    header_fill = PatternFill(start_color="01AB7B", end_color="01AB7B", fill_type="solid")
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    insumos = Insumo.objects.all()
    for insumo in insumos:
        ws.append([
            insumo.id,
            insumo.nombre,
            insumo.descripcion,
            insumo.proveedor.nombre,
            insumo.cantidad,
        ])

    # Establecer el ancho de las columnas
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15

    # Aplicar estilos a las celdas
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

    # Ajustar el ancho de la fila para el título
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 30

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_insumos.xlsx"'
    
    wb.save(response)
    return response

def reporte_insumo_pdf(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=landscape(letter))
    width, height = landscape(letter)

    margin = 40
    table_width = width - 2 * margin
    y_position = height - margin - 110
     # Ruta de la marca de agua
    watermark_path = finders.find('imagen/logo.png')
    p.saveState()

    # Ajustar transparencia y tamaño de la marca de agua
    p.setFillColor(colors.Color(1, 1, 1, alpha=0.1))  # Más transparente
    p.setStrokeColor(colors.Color(1, 1, 1, alpha=0.1))
    p.drawImage(watermark_path, x=(width - 600) / 2, y=(height - 600) / 2, width=600, height=600, mask='auto')

    p.restoreState()

    p.setFont("Helvetica-Bold", 24)
    p.drawCentredString(width / 2, y_position + 50, "LACTEOS HEDYBED")
    p.setFont("Helvetica", 18)
    p.drawCentredString(width / 2, y_position + 20, "Reporte de Insumos")

    data = [["ID", "Nombre", "Descripción", "Proveedor", "Cantidad"]]
    insumos = Insumo.objects.all()
    for insumo in insumos:
        data.append([
            insumo.id,
            insumo.nombre,
            insumo.descripcion,
            insumo.proveedor.nombre,
            insumo.cantidad,
        ])

    table = Table(data, colWidths=[1 * inch, 2* inch, 3 * inch, 2 * inch, 1.5 * inch, 2 * inch])  # Anchos ajustados

    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#01AB7B")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    table.setStyle(style)

    table.wrapOn(p, table_width, height - 2 * margin)
    table_x = margin
    table_y = y_position - len(data) * 30
    table.drawOn(p, table_x, table_y)

    p.showPage()
    p.save()
    
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte_insumos.pdf"'

    return response
def reporte_cliente_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"

    logo_path = finders.find('imagen/logo.png')
    img = Image(logo_path)
    img.height = 60
    img.width = 80
    ws.add_image(img, 'A1')

    ws.merge_cells('B1:D1')
    ws['B1'] = "LACTEOS HEDYBED"
    ws['B1'].font = Font(size=24, bold=True)
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:E2')
    ws['A2'] = "Clientes Registrados"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    headers = ["ID", "Nombre", "Tipo de Cliente", "Email", "Teléfono", "Dirección"]
    ws.append(headers)

    header_fill = PatternFill(start_color="01AB7B", end_color="01AB7B", fill_type="solid")
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    clientes = Cliente.objects.all()
    for cliente in clientes:
        ws.append([
            cliente.id,
            cliente.nombre,
            cliente.tipo_cliente,
            cliente.email,
            cliente.telefono,
            cliente.direccion,
        ])

    # Establecer el ancho de las columnas
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 30

    # Aplicar estilos a las celdas
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

    # Ajustar el ancho de la fila para el título
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 30

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_clientes.xlsx"'
    
    wb.save(response)
    return response


def reporte_cliente_pdf(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=landscape(letter))
    width, height = landscape(letter)

    margin = 40
    table_width = width - 2 * margin
    y_position = height - margin - 110

    # Ruta de la marca de agua
    watermark_path = finders.find('imagen/logo.png')
    p.saveState()

    # Ajustar transparencia y tamaño de la marca de agua
    p.setFillColor(colors.Color(1, 1, 1, alpha=0.1))  # Más transparente
    p.setStrokeColor(colors.Color(1, 1, 1, alpha=0.1))
    p.drawImage(watermark_path, x=(width - 600) / 2, y=(height - 600) / 2, width=600, height=600, mask='auto')

    p.restoreState()

    p.setFont("Helvetica-Bold", 24)
    p.drawCentredString(width / 2, y_position + 50, "LACTEOS HEDYBED")
    p.setFont("Helvetica", 18)
    p.drawCentredString(width / 2, y_position + 20, "Reporte de Clientes")

    data = [["ID", "Nombre", "Tipo de Cliente", "Email", "Teléfono", "Dirección"]]
    clientes = Cliente.objects.all()
    for cliente in clientes:
        data.append([
            cliente.id,
            cliente.nombre,
            cliente.tipo_cliente,
            cliente.email,
            cliente.telefono,
            cliente.direccion,
        ])

    table = Table(data, colWidths=[1 * inch, 2 * inch, 1.5 * inch, 2 * inch, 1.5 * inch, 2 * inch])

    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#01AB7B")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    table.setStyle(style)

    table.wrapOn(p, table_width, height - 2 * margin)
    table_x = margin
    table_y = y_position - len(data) * 30
    table.drawOn(p, table_x, table_y)

    p.showPage()
    p.save()
    
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte_clientes.pdf"'

    return response


def reporte_producto_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    logo_path = finders.find('imagen/logo.png')
    img = Image(logo_path)
    img.height = 60
    img.width = 80
    ws.add_image(img, 'A1')

    ws.merge_cells('B1:D1')
    ws['B1'] = "LACTEOS HEDYBED"
    ws['B1'].font = Font(size=24, bold=True)
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:E2')
    ws['A2'] = "Productos Registrados"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    headers = ["Código", "Nombre", "Presentación", "Unidad de Medida", "Cantidad", "Precio"]
    ws.append(headers)

    header_fill = PatternFill(start_color="01AB7B", end_color="01AB7B", fill_type="solid")
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    productos = Producto.objects.all()
    for producto in productos:
        ws.append([
            producto.codigo,
            producto.nombre,
            producto.presentacion,
            producto.unidad_medida,
            producto.cantidad,
            producto.formatear_precio(),  # Formateamos el precio
        ])

    # Establecer el ancho de las columnas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    # Aplicar estilos a las celdas
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=6):  # Cambiado max_col a 6
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

    # Ajustar el ancho de la fila para el título
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 30

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_productos.xlsx"'
    
    wb.save(response)
    return response
def reporte_producto_pdf(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=landscape(letter))
    width, height = landscape(letter)

    margin = 40
    table_width = width - 2 * margin
    y_position = height - margin - 110

    # Ruta de la marca de agua
    watermark_path = finders.find('imagen/logo.png')
    p.saveState()

    # Ajustar transparencia y tamaño de la marca de agua
    p.setFillColor(colors.Color(1, 1, 1, alpha=0.1))  # Más transparente
    p.setStrokeColor(colors.Color(1, 1, 1, alpha=0.1))
    p.drawImage(watermark_path, x=(width - 600) / 2, y=(height - 600) / 2, width=600, height=600, mask='auto')

    p.restoreState()

    p.setFont("Helvetica-Bold", 24)
    p.drawCentredString(width / 2, y_position + 50, "LACTEOS HEDYBED")
    p.setFont("Helvetica", 18)
    p.drawCentredString(width / 2, y_position + 20, "Reporte de Productos")

    data = [["Código", "Nombre", "Presentación", "Unidad de Medida", "Cantidad", "Precio"]]
    productos = Producto.objects.all()
    for producto in productos:
        data.append([
            producto.codigo,
            producto.nombre,
            producto.presentacion,
            producto.unidad_medida,
            producto.cantidad,
            producto.formatear_precio(), 
        ])

    table = Table(data, colWidths=[2 * inch, 2 * inch, 1.5 * inch, 1.5 * inch, 1 * inch, 2 * inch])

    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#01AB7B")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    table.setStyle(style)

    table.wrapOn(p, table_width, height - 2 * margin)
    table_x = margin
    table_y = y_position - len(data) * 30
    table.drawOn(p, table_x, table_y)

    p.showPage()
    p.save()
    
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte_productos.pdf"'

    return response

def reporte_movimiento_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transacciones"

    logo_path = finders.find('imagen/logo.png')
    img = Image(logo_path)
    img.height = 60
    img.width = 80
    ws.add_image(img, 'A1')

    ws.merge_cells('B1:D1')
    ws['B1'] = "LACTEOS HEDYBED"
    ws['B1'].font = Font(size=24, bold=True)
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:E2')
    ws['A2'] = "Transacciones Registradas"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    headers = ["Tipo", "Cliente", "Descripción", "Monto", "Fecha"]
    ws.append(headers)

    header_fill = PatternFill(start_color="01AB7B", end_color="01AB7B", fill_type="solid")
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    transacciones = Transaccion.objects.all()
    for transaccion in transacciones:
        ws.append([
            transaccion.tipo,
            str(transaccion.cliente),  # Asegúrate de que el método str esté definido en Cliente
            transaccion.descripcion,
            transaccion.monto,
            transaccion.fecha,
        ])

    # Establecer el ancho de las columnas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15

    # Aplicar estilos a las celdas
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

    # Ajustar el ancho de la fila para el título
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 30

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_transacciones.xlsx"'
    
    wb.save(response)
    return response

def reporte_movimiento_pdf(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=landscape(letter))
    width, height = landscape(letter)

    margin = 40
    table_width = width - 2 * margin
    y_position = height - margin - 110

    # Ruta de la marca de agua
    watermark_path = finders.find('imagen/logo.png')
    p.saveState()

    # Ajustar transparencia y tamaño de la marca de agua
    p.setFillColor(colors.Color(1, 1, 1, alpha=0.1))  # Más transparente
    p.setStrokeColor(colors.Color(1, 1, 1, alpha=0.1))
    p.drawImage(watermark_path, x=(width - 600) / 2, y=(height - 600) / 2, width=600, height=600, mask='auto')

    p.restoreState()

    p.setFont("Helvetica-Bold", 24)
    p.drawCentredString(width / 2, y_position + 50, "LACTEOS HEDYBED")
    p.setFont("Helvetica", 18)
    p.drawCentredString(width / 2, y_position + 20, "Reporte de Transacciones")

    data = [["Tipo", "Cliente", "Descripción", "Monto", "Fecha"]]
    transacciones = Transaccion.objects.all()
    for transaccion in transacciones:
        data.append([
            transaccion.tipo,
            transaccion.cliente,  # Asegúrate de que el método str esté definido en Cliente
            transaccion.descripcion,
            transaccion.monto,
            transaccion.fecha,
        ])

    table = Table(data, colWidths=[1.5 * inch, 2 * inch, 3 * inch, 2 * inch, 1.5 * inch])

    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#01AB7B")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    table.setStyle(style)

    table.wrapOn(p, table_width, height - 2 * margin)
    table_x = margin
    table_y = y_position - len(data) * 30
    table.drawOn(p, table_x, table_y)

    p.showPage()
    p.save()
    
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte_transacciones.pdf"'

    return response